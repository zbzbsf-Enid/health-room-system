import io
import pandas as pd
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from openpyxl.utils import get_column_letter

def generate_monthly_report_excel(df):
    """
    接收 Streamlit 內部的 DataFrame 資料，轉化為 115 學年度標準 Excel 檔案 (BytesIO)
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "115年9月用藥月報表"
    ws.views.sheetView[0].showGridLines = True

    # 樣式設定
    font_title = Font(name="微軟正黑體", size=14, bold=True, color="1F4E78")
    font_header_white = Font(name="微軟正黑體", size=10, bold=True, color="FFFFFF")
    font_header_dark = Font(name="微軟正黑體", size=9, bold=True, color="333333")
    font_data = Font(name="微軟正黑體", size=9)
    font_red_bold = Font(name="微軟正黑體", size=9, bold=True, color="C00000")
    font_blue_bold = Font(name="微軟正黑體", size=9, bold=True, color="002060")
    font_orange_bold = Font(name="微軟正黑體", size=9, bold=True, color="C65911")

    fill_blue = PatternFill(start_color="1F4E78", fill_type="solid")
    fill_green = PatternFill(start_color="2E75B6", fill_type="solid")
    fill_orange = PatternFill(start_color="C65911", fill_type="solid")
    fill_gray = PatternFill(start_color="F2F2F2", fill_type="solid")
    fill_light_yellow = PatternFill(start_color="FFF2CC", fill_type="solid")
    fill_light_orange = PatternFill(start_color="FCE4D6", fill_type="solid")

    thin_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='thin', color='D9D9D9'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))
    thick_right_border = Border(left=Side(style='thin', color='D9D9D9'), right=Side(style='medium', color='1F4E78'), top=Side(style='thin', color='D9D9D9'), bottom=Side(style='thin', color='D9D9D9'))

    align_center = Alignment(horizontal='center', vertical='center', wrap_text=True)
    align_left = Alignment(horizontal='left', vertical='center')
    align_right = Alignment(horizontal='right', vertical='center')

    # 標題
    ws.merge_cells('A1:AK1')
    ws['A1'] = "國立臺北大學衛保組 115學年度上學期藥品使用月報與全學期統計表 (115年9月起)"
    ws['A1'].font, ws['A1'].alignment = font_title, Alignment(horizontal='left', vertical='center')
    ws.row_dimensions[1].height = 35

    sept_dates = ['9/1', '9/2', '9/3', '9/4', '9/7', '9/8', '9/9', '9/10', '9/11', '9/14', '9/15', '9/16', '9/17', '9/18', '9/21', '9/22', '9/23', '9/24', '9/25', '9/28', '9/29', '9/30']
    headers = ["藥品名稱\n(商品名/中文)", "115年8月\n剩餘量"] + sept_dates + [
        "當月使用\n總量", "購入量", "過期報銷", "公藥使用", "115年9月\n期末剩餘量", "實體盤點\n數量", "有效期限",
        "9月\n消耗量", "10月\n消耗量", "11月\n消耗量", "12月\n消耗量", "1月\n消耗量", "全學期\n使用總量"
    ]

    ws.row_dimensions[2].height = 32
    daily_end_idx = 24

    for idx, header in enumerate(headers, 1):
        cell = ws.cell(row=2, column=idx, value=header)
        cell.alignment, cell.border = align_center, thin_border
        if idx in [1, 2]:
            cell.fill, cell.font = fill_blue, font_header_white
        elif 3 <= idx <= daily_end_idx:
            cell.fill, cell.font = fill_gray, font_header_dark
        elif daily_end_idx + 1 <= idx <= daily_end_idx + 7:
            cell.fill, cell.font = fill_green, font_header_white
        else:
            cell.fill, cell.font = fill_orange, font_header_white

    # 填入資料與公式
    for i, row in df.iterrows():
        r = i + 3
        ws.row_dimensions[r].height = 22

        name_val = str(row['name']) if 'name' in row and pd.notna(row['name']) else str(row.get('藥品名稱', ''))
        cname_val = str(row['chinese_name']) if 'chinese_name' in row and pd.notna(row['chinese_name']) else ""
        full_name = f"{name_val} ({cname_val})" if cname_val and cname_val != 'nan' else name_val
        
        ws.cell(row=r, column=1, value=full_name).alignment = align_left
        ws.cell(row=r, column=1).font, ws.cell(row=r, column=1).border = font_data, thin_border

        stock_aug = row['aug_stock'] if 'aug_stock' in row and pd.notna(row['aug_stock']) else row.get('stock', 0)
        ws.cell(row=r, column=2, value=stock_aug).font = font_red_bold
        ws.cell(row=r, column=2).alignment, ws.cell(row=r, column=2).border = align_right, thin_border

        for d_idx, day_str in enumerate(sept_dates, start=3):
            used_val = row[day_str] if day_str in row and pd.notna(row[day_str]) and row[day_str] > 0 else ""
            c = ws.cell(row=r, column=d_idx, value=used_val)
            c.border, c.alignment, c.font = thin_border, align_center, font_data

        ws.cell(row=r, column=25, value=f"=SUM(C{r}:X{r})").font = font_blue_bold
        ws.cell(row=r, column=25).fill = fill_light_yellow
        ws.cell(row=r, column=25).alignment, ws.cell(row=r, column=25).border = align_right, thin_border

        ws.cell(row=r, column=26, value=row.get('purchased', 0)).alignment = align_right
        ws.cell(row=r, column=26).border = thin_border
        ws.cell(row=r, column=27, value=row.get('expired', 0)).alignment = align_right
        ws.cell(row=r, column=27).border = thin_border
        ws.cell(row=r, column=28, value=row.get('public_use', 0)).alignment = align_right
        ws.cell(row=r, column=28).border = thin_border

        ws.cell(row=r, column=29, value=f"=B{r}+Z{r}-Y{r}-AA{r}-AB{r}").font = font_red_bold
        ws.cell(row=r, column=29).alignment, ws.cell(row=r, column=29).border = align_right, thin_border

        ws.cell(row=r, column=30, value=f"=AC{r}").alignment = align_right
        ws.cell(row=r, column=30).border = thin_border

        expiry_val = str(row['expiry']) if 'expiry' in row and pd.notna(row['expiry']) else ""
        ws.cell(row=r, column=31, value=expiry_val).alignment = align_center
        ws.cell(row=r, column=31).border = thick_right_border

        ws.cell(row=r, column=32, value=f"=Y{r}").alignment = align_right
        ws.cell(row=r, column=32).border = thin_border
        for c_idx in range(33, 37):
            ws.cell(row=r, column=c_idx, value=0).alignment = align_right
            ws.cell(row=r, column=c_idx).border = thin_border

        ws.cell(row=r, column=37, value=f"=SUM(AF{r}:AJ{r})").font = font_orange_bold
        ws.cell(row=r, column=37).fill = fill_light_orange
        ws.cell(row=r, column=37).alignment, ws.cell(row=r, column=37).border = align_right, thin_border

    ws.column_dimensions['A'].width = 28
    ws.column_dimensions['B'].width = 12
    for c in range(3, daily_end_idx + 1):
        ws.column_dimensions[get_column_letter(c)].width = 5.5
    ws.column_dimensions['Y'].width = 12
    ws.column_dimensions['AC'].width = 12

    # 轉為 Bytes 輸出給 Streamlit 網頁下載
    output = io.BytesIO()
    wb.save(output)
    output.seek(0)
    return output
